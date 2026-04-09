import os
import re
import io
import csv
from datetime import datetime, timedelta

from flask import Flask, request, jsonify, render_template, send_file
from werkzeug.utils import secure_filename

from models import db, Contrato, Conta, Comparacao, Importacao, Configuracao, PlanoPreco, FaturaLinha, CadastroLinha, CoopernacVoz, CoopernacDados, CoopernacResumo
from pdf_extractor import PDFExtractor
from comparator import Comparator

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ============================================================
# Configuração da aplicação
# ============================================================
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
IS_VERCEL = os.environ.get('VERCEL', False)
IS_RENDER = os.environ.get('RENDER', False)

# Na Vercel, o filesystem é read-only exceto /tmp
# No Render, usamos disco persistente montado em /opt/render/project/data
if IS_VERCEL or IS_RENDER:
    DB_DIR = '/tmp'
    UPLOAD_DIR = '/tmp/uploads'
else:
    DB_DIR = BASE_DIR
    UPLOAD_DIR = os.path.join(BASE_DIR, 'uploads')

app = Flask(__name__)
app.config['SQLALCHEMY_DATABASE_URI'] = f'sqlite:///{os.path.join(DB_DIR, "auditoria.db")}'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['UPLOAD_FOLDER'] = UPLOAD_DIR
if IS_VERCEL:
    app.config['MAX_CONTENT_LENGTH'] = 4 * 1024 * 1024    # 4 MB (limite Vercel ~4.5 MB)
else:
    app.config['MAX_CONTENT_LENGTH'] = 100 * 1024 * 1024  # 100 MB
app.config['SECRET_KEY'] = os.environ.get('SECRET_KEY', os.urandom(32).hex())

db.init_app(app)
ALLOWED_EXTENSIONS = {'pdf'}

def allowed_file(filename: str) -> bool:
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


def init_db():
    """Inicializa banco de dados e configurações padrão."""
    with app.app_context():
        db.create_all()
        os.makedirs(os.path.join(app.config['UPLOAD_FOLDER'], 'contratos'), exist_ok=True)
        os.makedirs(os.path.join(app.config['UPLOAD_FOLDER'], 'contas'), exist_ok=True)

        defaults = [
            ('tolerancia_tipo', 'percentual', 'Tipo de tolerância: fixo (R$) ou percentual (%)'),
            ('tolerancia_valor', '5.0', 'Valor da tolerância'),
            ('apenas_recorrentes', 'false', 'Considerar apenas cobranças recorrentes'),
            ('ignorar_extras', 'false', 'Ignorar taxas extras e multas'),
            ('empresa_nome', 'Auditoria Telecom', 'Nome da empresa'),
        ]
        for chave, valor, descricao in defaults:
            if not Configuracao.query.filter_by(chave=chave).first():
                db.session.add(Configuracao(chave=chave, valor=valor, descricao=descricao))
        db.session.commit()


# Em ambientes cloud, inicializar o DB automaticamente
if IS_VERCEL or IS_RENDER:
    _cloud_db_initialized = False

    @app.before_request
    def _ensure_db():
        global _cloud_db_initialized
        if not _cloud_db_initialized:
            init_db()
            _cloud_db_initialized = True


@app.errorhandler(413)
def request_entity_too_large(error):
    limit = '4 MB (limite da Vercel)' if IS_VERCEL else '100 MB'
    return jsonify({'error': f'Arquivo muito grande. Limite máximo: {limit}'}), 413


# ============================================================
# Página principal
# ============================================================
@app.route('/')
def index():
    return render_template('index.html')


# ============================================================
# DASHBOARD
# ============================================================
@app.route('/api/dashboard')
def dashboard():
    total_contratos = Contrato.query.filter_by(ativo=True).count()
    total_linhas_ativas = db.session.query(
        db.func.count(db.func.distinct(Contrato.linha_telefone))
    ).filter(Contrato.ativo == True).scalar() or 0
    total_contas = Conta.query.count()
    total_importacoes = Importacao.query.count()

    # Contagens por status
    ok_count = Comparacao.query.filter_by(status_comparacao='ok').count()
    aprox_count = Comparacao.query.filter_by(status_comparacao='aproximado').count()
    div_count = Comparacao.query.filter_by(status_comparacao='divergente').count()
    sem_contrato_count = Comparacao.query.filter_by(status_comparacao='sem_contrato').count()
    sem_fatura_count = Comparacao.query.filter_by(status_comparacao='sem_fatura').count()
    ambiguo_count = Comparacao.query.filter_by(status_comparacao='ambiguo').count()
    total_comparacoes = Comparacao.query.count()

    conformidade = 0
    if total_comparacoes > 0:
        conformidade = round((ok_count + aprox_count) / total_comparacoes * 100, 1)

    # Valores financeiros
    total_contratado = db.session.query(
        db.func.sum(Comparacao.valor_contratado)
    ).filter(Comparacao.valor_contratado.isnot(None)).scalar() or 0

    total_faturado = db.session.query(
        db.func.sum(Comparacao.valor_fatura)
    ).filter(Comparacao.valor_fatura.isnot(None)).scalar() or 0

    # Evolução mensal (últimas 6 competências)
    monthly_raw = db.session.query(
        Comparacao.competencia,
        db.func.count(Comparacao.id).label('total'),
        db.func.sum(db.case((Comparacao.status_comparacao == 'ok', 1), else_=0)).label('ok'),
        db.func.sum(db.case((Comparacao.status_comparacao == 'divergente', 1), else_=0)).label('divergente'),
        db.func.sum(db.case((Comparacao.status_comparacao == 'aproximado', 1), else_=0)).label('aproximado'),
    ).filter(
        Comparacao.competencia.isnot(None)
    ).group_by(Comparacao.competencia).order_by(
        Comparacao.competencia.desc()
    ).limit(6).all()

    monthly_data = [
        {
            'competencia': r.competencia,
            'total': r.total,
            'ok': r.ok or 0,
            'divergente': r.divergente or 0,
            'aproximado': r.aproximado or 0,
        }
        for r in reversed(monthly_raw)
    ]

    # Por operadora
    op_raw = db.session.query(
        Conta.operadora,
        db.func.count(Conta.id).label('count')
    ).group_by(Conta.operadora).all()
    operadoras_data = [
        {'operadora': r.operadora or 'Desconhecida', 'count': r.count}
        for r in op_raw
    ]

    # --- Dados de Faturas Importadas (FaturaLinha) ---
    fl_total = FaturaLinha.query.count()
    fl_total_faturado = db.session.query(
        db.func.sum(FaturaLinha.valor_fatura)
    ).filter(FaturaLinha.valor_fatura.isnot(None)).scalar() or 0
    fl_total_contrato = db.session.query(
        db.func.sum(FaturaLinha.valor_contrato)
    ).filter(FaturaLinha.valor_contrato.isnot(None)).scalar() or 0
    fl_diferenca = db.session.query(
        db.func.sum(FaturaLinha.diferenca)
    ).filter(FaturaLinha.diferenca.isnot(None)).scalar() or 0
    fl_divergentes = FaturaLinha.query.filter_by(status='divergente').count()
    fl_ok = FaturaLinha.query.filter_by(status='ok').count()
    total_planos = PlanoPreco.query.count()

    return jsonify({
        'total_contratos': total_contratos,
        'total_linhas_ativas': total_linhas_ativas,
        'total_contas': total_contas,
        'total_importacoes': total_importacoes,
        'total_comparacoes': total_comparacoes,
        'ok_count': ok_count,
        'aprox_count': aprox_count,
        'div_count': div_count,
        'sem_contrato_count': sem_contrato_count,
        'sem_fatura_count': sem_fatura_count,
        'ambiguo_count': ambiguo_count,
        'conformidade': conformidade,
        'total_contratado': round(float(total_contratado), 2),
        'total_faturado': round(float(total_faturado), 2),
        'monthly_data': monthly_data,
        'operadoras': operadoras_data,
        # Novos campos: faturas importadas
        'fl_total': fl_total,
        'fl_total_faturado': round(float(fl_total_faturado), 2),
        'fl_total_contrato': round(float(fl_total_contrato), 2),
        'fl_diferenca': round(float(fl_diferenca), 2),
        'fl_divergentes': fl_divergentes,
        'fl_ok': fl_ok,
        'total_planos': total_planos,
    })


# ============================================================
# CONTRATOS
# ============================================================
@app.route('/api/contratos', methods=['GET'])
def list_contratos():
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 50, type=int)
    search = request.args.get('search', '').strip()
    operadora = request.args.get('operadora', '').strip()
    ativo_param = request.args.get('ativo', 'true')

    query = Contrato.query
    if ativo_param == 'true':
        query = query.filter_by(ativo=True)
    elif ativo_param == 'false':
        query = query.filter_by(ativo=False)

    if search:
        like = f'%{search}%'
        query = query.filter(
            db.or_(
                Contrato.numero_contrato.ilike(like),
                Contrato.linha_telefone.ilike(like),
                Contrato.cliente.ilike(like),
            )
        )
    if operadora:
        query = query.filter(Contrato.operadora.ilike(f'%{operadora}%'))

    total = query.count()
    items = (query.order_by(Contrato.data_importacao.desc())
             .offset((page - 1) * per_page).limit(per_page).all())

    return jsonify({
        'data': [c.to_dict() for c in items],
        'total': total,
        'page': page,
        'per_page': per_page,
        'pages': (total + per_page - 1) // per_page,
    })


@app.route('/api/contratos/<int:cid>', methods=['GET'])
def get_contrato(cid):
    c = Contrato.query.get_or_404(cid)
    return jsonify(c.to_dict())


@app.route('/api/contratos', methods=['POST'])
def create_contrato():
    data = request.get_json()
    if not data:
        return jsonify({'error': 'Dados inválidos'}), 400
    if not data.get('linha_telefone') or data.get('valor_contratado') is None:
        return jsonify({'error': 'Linha e valor contratado são obrigatórios'}), 400

    c = Contrato(
        numero_contrato=data.get('numero_contrato', '').strip() or None,
        linha_telefone=re.sub(r'\D', '', str(data['linha_telefone'])),
        valor_contratado=float(str(data['valor_contratado']).replace(',', '.')),
        cliente=data.get('cliente', '').strip() or None,
        operadora=data.get('operadora', '').strip() or None,
        observacoes=data.get('observacoes', '').strip() or None,
        ativo=True,
    )
    db.session.add(c)
    db.session.commit()
    return jsonify(c.to_dict()), 201


@app.route('/api/contratos/<int:cid>', methods=['PUT'])
def update_contrato(cid):
    c = Contrato.query.get_or_404(cid)
    data = request.get_json()
    if not data:
        return jsonify({'error': 'Dados inválidos'}), 400

    if 'numero_contrato' in data:
        c.numero_contrato = data['numero_contrato']
    if 'linha_telefone' in data:
        c.linha_telefone = re.sub(r'\D', '', str(data['linha_telefone']))
    if 'valor_contratado' in data:
        c.valor_contratado = float(str(data['valor_contratado']).replace(',', '.'))
    if 'cliente' in data:
        c.cliente = data['cliente']
    if 'operadora' in data:
        c.operadora = data['operadora']
    if 'observacoes' in data:
        c.observacoes = data['observacoes']
    if 'ativo' in data:
        c.ativo = bool(data['ativo'])

    db.session.commit()
    return jsonify(c.to_dict())


@app.route('/api/contratos/<int:cid>', methods=['DELETE'])
def delete_contrato(cid):
    c = Contrato.query.get_or_404(cid)
    c.ativo = False  # soft delete
    db.session.commit()
    return jsonify({'message': 'Contrato desativado'})


@app.route('/api/contratos/upload-pdf', methods=['POST'])
def upload_contrato_pdf():
    if 'file' not in request.files:
        return jsonify({'error': 'Nenhum arquivo enviado'}), 400
    f = request.files['file']
    if not f.filename or not allowed_file(f.filename):
        return jsonify({'error': 'Apenas PDFs são aceitos'}), 400

    fname = secure_filename(f.filename)
    ts = datetime.now().strftime('%Y%m%d_%H%M%S_')
    save_name = ts + fname
    fpath = os.path.join(app.config['UPLOAD_FOLDER'], 'contratos', save_name)
    f.save(fpath)

    extractor = PDFExtractor()
    try:
        extracted = extractor.extract_from_contract_pdf(fpath)
        raw_text  = extractor.get_raw_text(fpath)
    except Exception as e:
        extracted = []
        raw_text  = ''
    return jsonify({'filename': save_name, 'extracted': extracted,
                    'count': len(extracted), 'raw_text': raw_text[:3000]})


@app.route('/api/contratos/salvar-lote', methods=['POST'])
def save_contratos_lote():
    data = request.get_json()
    if not data or 'contratos' not in data:
        return jsonify({'error': 'Dados inválidos'}), 400

    importacao = Importacao(
        tipo='contrato',
        arquivo_nome=data.get('arquivo_nome', 'importação manual'),
        status='concluido',
        total_registros=len(data['contratos'])
    )
    db.session.add(importacao)
    db.session.flush()

    saved, errors = 0, []
    for row in data['contratos']:
        if not row.get('linha_telefone') or row.get('valor_contratado') is None:
            errors.append(f"Dados incompletos: {row.get('linha_telefone', '?')}")
            continue
        try:
            valor = float(str(row['valor_contratado']).replace(',', '.'))
            c = Contrato(
                numero_contrato=row.get('numero_contrato') or None,
                linha_telefone=re.sub(r'\D', '', str(row['linha_telefone'])),
                valor_contratado=valor,
                cliente=row.get('cliente') or None,
                operadora=row.get('operadora') or None,
                arquivo_pdf_origem=row.get('arquivo_pdf_origem'),
                importacao_id=importacao.id,
                ativo=True,
            )
            db.session.add(c)
            saved += 1
        except Exception as e:
            errors.append(str(e))

    db.session.commit()
    return jsonify({'saved': saved, 'errors': errors, 'importacao_id': importacao.id})


# ============================================================
# CONTAS
# ============================================================
@app.route('/api/contas', methods=['GET'])
def list_contas():
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 50, type=int)
    competencia = request.args.get('competencia', '').strip()

    query = Conta.query
    if competencia:
        query = query.filter_by(competencia=competencia)

    total = query.count()
    items = (query.order_by(Conta.data_importacao.desc())
             .offset((page - 1) * per_page).limit(per_page).all())

    return jsonify({
        'data': [c.to_dict() for c in items],
        'total': total,
        'page': page,
        'per_page': per_page,
    })


@app.route('/api/contas/upload-pdf', methods=['POST'])
def upload_conta_pdf():
    if 'file' not in request.files:
        return jsonify({'error': 'Nenhum arquivo enviado'}), 400
    f = request.files['file']
    if not f.filename or not allowed_file(f.filename):
        return jsonify({'error': 'Apenas PDFs são aceitos'}), 400

    fname = secure_filename(f.filename)
    ts = datetime.now().strftime('%Y%m%d_%H%M%S_')
    save_name = ts + fname
    fpath = os.path.join(app.config['UPLOAD_FOLDER'], 'contas', save_name)
    f.save(fpath)

    extractor = PDFExtractor()
    try:
        extracted = extractor.extract_from_invoice_pdf(fpath)
        raw_text  = extractor.get_raw_text(fpath)
    except Exception as e:
        extracted = []
        raw_text  = ''
    return jsonify({'filename': save_name, 'extracted': extracted,
                    'count': len(extracted), 'raw_text': raw_text[:3000]})


@app.route('/api/contas/salvar-lote', methods=['POST'])
def save_contas_lote():
    data = request.get_json()
    if not data or 'contas' not in data:
        return jsonify({'error': 'Dados inválidos'}), 400

    importacao = Importacao(
        tipo='conta',
        arquivo_nome=data.get('arquivo_nome', 'importação manual'),
        status='concluido',
        total_registros=len(data['contas'])
    )
    db.session.add(importacao)
    db.session.flush()

    saved, errors = 0, []
    competencia_ref = None
    for row in data['contas']:
        if not row.get('linha_telefone') or row.get('valor_fatura') is None:
            errors.append(f"Dados incompletos: {row.get('linha_telefone', '?')}")
            continue
        try:
            valor = float(str(row['valor_fatura']).replace(',', '.'))
            c = Conta(
                linha_telefone=re.sub(r'\D', '', str(row['linha_telefone'])),
                valor_fatura=valor,
                competencia=row.get('competencia') or None,
                operadora=row.get('operadora') or None,
                numero_fatura=row.get('numero_fatura') or None,
                arquivo_pdf_origem=row.get('arquivo_pdf_origem'),
                importacao_id=importacao.id,
            )
            db.session.add(c)
            saved += 1
            if row.get('competencia') and not competencia_ref:
                competencia_ref = row['competencia']
        except Exception as e:
            errors.append(str(e))

    db.session.commit()

    # Executar comparação automática
    comp_result = {}
    if saved > 0:
        try:
            comparator = Comparator()
            comp_result = comparator.run_comparison(
                competencia=competencia_ref,
                importacao_ids=[importacao.id]
            )
        except Exception as e:
            comp_result = {'error': str(e)}

    return jsonify({
        'saved': saved,
        'errors': errors,
        'importacao_id': importacao.id,
        'comparacao': comp_result.get('totais', {}),
    })


# ============================================================
# COMPARAÇÕES
# ============================================================
@app.route('/api/comparacoes', methods=['GET'])
def list_comparacoes():
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 100, type=int)
    status = request.args.get('status', '').strip()
    competencia = request.args.get('competencia', '').strip()
    search = request.args.get('search', '').strip()
    operadora = request.args.get('operadora', '').strip()

    query = Comparacao.query
    if status:
        query = query.filter_by(status_comparacao=status)
    if competencia:
        query = query.filter_by(competencia=competencia)
    if search:
        like = f'%{search}%'
        query = query.filter(
            db.or_(
                Comparacao.linha_telefone.ilike(like),
                Comparacao.numero_contrato.ilike(like),
            )
        )
    if operadora:
        query = query.filter(Comparacao.operadora.ilike(f'%{operadora}%'))

    total = query.count()
    items = (query.order_by(Comparacao.data_processamento.desc())
             .offset((page - 1) * per_page).limit(per_page).all())

    return jsonify({
        'data': [c.to_dict() for c in items],
        'total': total,
        'page': page,
        'per_page': per_page,
    })


@app.route('/api/comparar', methods=['POST'])
def run_comparison():
    data = request.get_json() or {}
    competencia = data.get('competencia')
    importacao_ids = data.get('importacao_ids')

    comparator = Comparator()
    result = comparator.run_comparison(competencia=competencia, importacao_ids=importacao_ids)
    return jsonify(result)


@app.route('/api/comparacoes/<int:cid>/historico', methods=['GET'])
def get_historico_linha(cid):
    comp = Comparacao.query.get_or_404(cid)
    historico = (Comparacao.query
                 .filter_by(linha_telefone=comp.linha_telefone)
                 .order_by(Comparacao.data_processamento.desc())
                 .limit(12).all())
    return jsonify([h.to_dict() for h in historico])


# ============================================================
# HISTÓRICO DE IMPORTAÇÕES
# ============================================================
@app.route('/api/historico', methods=['GET'])
def list_historico():
    importacoes = (Importacao.query
                   .order_by(Importacao.data_importacao.desc())
                   .limit(200).all())
    return jsonify([i.to_dict() for i in importacoes])


@app.route('/api/historico/<int:iid>', methods=['DELETE'])
def delete_importacao(iid):
    importacao = Importacao.query.get_or_404(iid)
    # Remove comparações relacionadas
    if importacao.tipo == 'conta':
        for conta in importacao.contas:
            Comparacao.query.filter_by(conta_id=conta.id).delete()
    db.session.delete(importacao)
    db.session.commit()
    return jsonify({'message': 'Importação removida'})


# ============================================================
# CONFIGURAÇÕES
# ============================================================
@app.route('/api/configuracoes', methods=['GET'])
def get_configuracoes():
    configs = Configuracao.query.all()
    return jsonify({c.chave: c.valor for c in configs})


@app.route('/api/configuracoes', methods=['PUT'])
def update_configuracoes():
    data = request.get_json()
    if not data:
        return jsonify({'error': 'Dados inválidos'}), 400
    for chave, valor in data.items():
        cfg = Configuracao.query.filter_by(chave=chave).first()
        if cfg:
            cfg.valor = str(valor)
        else:
            db.session.add(Configuracao(chave=chave, valor=str(valor)))
    db.session.commit()
    return jsonify({'message': 'Configurações salvas'})


# ============================================================
# EXPORTAÇÃO
# ============================================================
def _get_comparacoes_for_export(competencia, status_filter):
    query = Comparacao.query
    if competencia:
        query = query.filter_by(competencia=competencia)
    if status_filter:
        query = query.filter_by(status_comparacao=status_filter)
    return query.order_by(Comparacao.competencia, Comparacao.status_comparacao).all()


STATUS_LABELS = {
    'ok': 'Conforme',
    'aproximado': 'Aproximado',
    'divergente': 'Divergente',
    'sem_contrato': 'Sem Contrato',
    'sem_fatura': 'Sem Fatura',
    'ambiguo': 'Ambíguo',
}

STATUS_COLORS_XLSX = {
    'ok': '92D050',       # verde
    'aproximado': 'FFD966',  # amarelo
    'divergente': 'FF7070',  # vermelho
    'sem_contrato': 'BDD7EE',  # azul claro
    'sem_fatura': 'D9D9D9',    # cinza
    'ambiguo': 'F4A460',       # laranja
}


@app.route('/api/exportar/excel')
def export_excel():
    competencia = request.args.get('competencia', '').strip()
    status_filter = request.args.get('status', '').strip()
    comparacoes = _get_comparacoes_for_export(competencia, status_filter)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Auditoria Telecom'
    ws.freeze_panes = 'A2'

    headers = [
        'Status', 'Linha Telefone', 'Nº Contrato', 'Valor Contratado (R$)',
        'Valor Fatura (R$)', 'Diferença (R$)', 'Diferença (%)',
        'Competência', 'Operadora', 'Observação', 'Data Processamento'
    ]

    # Cabeçalho
    header_fill = PatternFill('solid', fgColor='1F3864')
    header_font = Font(color='FFFFFF', bold=True, size=10)
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Dados
    for row_idx, comp in enumerate(comparacoes, 2):
        row_color = STATUS_COLORS_XLSX.get(comp.status_comparacao, 'FFFFFF')
        fill = PatternFill('solid', fgColor=row_color)

        row_data = [
            STATUS_LABELS.get(comp.status_comparacao, comp.status_comparacao),
            comp.linha_telefone or '',
            comp.numero_contrato or '',
            comp.valor_contratado,
            comp.valor_fatura,
            comp.diferenca_valor,
            f'{comp.diferenca_percentual:.1f}%' if comp.diferenca_percentual is not None else '',
            comp.competencia or '',
            comp.operadora or '',
            comp.observacao or '',
            comp.data_processamento.strftime('%d/%m/%Y %H:%M') if comp.data_processamento else '',
        ]

        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.fill = fill
            cell.font = Font(size=9)

    # Ajustar largura das colunas
    col_widths = [14, 18, 15, 20, 18, 16, 14, 13, 14, 45, 18]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    fname = f'auditoria_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=fname,
    )


# ============================================================
# FATURA PDF → XLS
# ============================================================

def _build_fatura_xls(linhas: list, meta: dict) -> io.BytesIO:
    """Gera um arquivo XLS a partir das linhas extraídas de uma fatura."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Fatura Vivo'
    ws.freeze_panes = 'A2'

    headers = ['Número Vivo', 'Plano', 'Valor Total R$']
    header_fill = PatternFill('solid', fgColor='660099')   # roxo Vivo
    header_font = Font(color='FFFFFF', bold=True, size=10)
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 18

    alt_fill = PatternFill('solid', fgColor='F3E8FF')
    for ri, item in enumerate(linhas, 2):
        # Formata número como XX-XXXXX-XXXX
        num = item.get('numero_vivo', '')
        if re.match(r'^\d{10,11}$', num):
            if len(num) == 11:
                num = f'{num[:2]}-{num[2:7]}-{num[7:]}'
            else:
                num = f'{num[:2]}-{num[2:6]}-{num[6:]}'
        ws.cell(row=ri, column=1, value=num).font = Font(size=9)
        plano_cell = ws.cell(row=ri, column=2, value=item.get('plano', ''))
        plano_cell.font = Font(size=9)
        valor = item.get('valor_total')
        valor_cell = ws.cell(row=ri, column=3, value=valor)
        valor_cell.font = Font(size=9)
        valor_cell.number_format = '#,##0.00'
        valor_cell.alignment = Alignment(horizontal='right')
        if ri % 2 == 0:
            for c in range(1, 4):
                ws.cell(row=ri, column=c).fill = alt_fill

    # Linha de total
    if linhas:
        total_row = len(linhas) + 2
        ws.cell(row=total_row, column=2, value='TOTAL').font = Font(bold=True, size=9)
        ws.cell(row=total_row, column=3,
                value=f'=SUM(C2:C{total_row-1})').font = Font(bold=True, size=9)
        ws.cell(row=total_row, column=3).number_format = '#,##0.00'

    # Metadados em aba separada
    ws_meta = wb.create_sheet('Info')
    ws_meta.append(['Campo', 'Valor'])
    ws_meta.append(['Arquivo', meta.get('filename', '')])
    ws_meta.append(['Competência', meta.get('competencia', '')])
    ws_meta.append(['Operadora', meta.get('operadora', '')])
    ws_meta.append(['Nº Fatura', meta.get('numero_fatura', '')])
    ws_meta.append(['Linhas extraídas', len(linhas)])

    col_widths = [20, 45, 18]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


@app.route('/api/faturas/pdf-preview', methods=['POST'])
def fatura_pdf_preview():
    """Extrai dados da fatura e retorna JSON para preview."""
    if 'file' not in request.files:
        return jsonify({'error': 'Nenhum arquivo enviado'}), 400
    f = request.files['file']
    if not f.filename or not allowed_file(f.filename):
        return jsonify({'error': 'Arquivo inválido. Somente PDF.'}), 400

    filename = datetime.now().strftime('%Y%m%d_%H%M%S_') + secure_filename(f.filename)
    path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
    os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
    f.save(path)

    try:
        extractor = PDFExtractor()
        data = extractor.extract_fatura_vivo(path)
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        try:
            os.remove(path)
        except Exception:
            pass

    return jsonify(data)


@app.route('/api/faturas/pdf-para-xls', methods=['POST'])
def fatura_pdf_para_xls():
    """Converte fatura PDF para XLS e retorna para download."""
    if 'file' not in request.files:
        return jsonify({'error': 'Nenhum arquivo enviado'}), 400
    f = request.files['file']
    if not f.filename or not allowed_file(f.filename):
        return jsonify({'error': 'Arquivo inválido. Somente PDF.'}), 400

    filename = datetime.now().strftime('%Y%m%d_%H%M%S_') + secure_filename(f.filename)
    path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
    os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
    f.save(path)

    try:
        extractor = PDFExtractor()
        data = extractor.extract_fatura_vivo(path)
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        try:
            os.remove(path)
        except Exception:
            pass

    xls_buf = _build_fatura_xls(data.get('linhas', []), data)
    base = os.path.splitext(f.filename)[0]
    dl_name = f'{base}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    return send_file(
        xls_buf,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=dl_name,
    )


@app.route('/api/faturas/xls-from-json', methods=['POST'])
def fatura_xls_from_json():
    """Gera XLS a partir de JSON com linhas editadas no frontend."""
    data = request.get_json(force=True, silent=True) or {}
    linhas = data.get('linhas', [])
    meta   = data.get('meta', {})
    if not isinstance(linhas, list):
        return jsonify({'error': 'Formato inválido'}), 400
    xls_buf = _build_fatura_xls(linhas, meta)
    base = re.sub(r'\.pdf$', '', meta.get('filename', 'fatura'), flags=re.IGNORECASE)
    dl_name = f'{base}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    return send_file(
        xls_buf,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=dl_name,
    )


# ============================================================
# PLANOS / PREÇOS
# ============================================================
@app.route('/api/planos', methods=['GET'])
def list_planos():
    planos = PlanoPreco.query.order_by(PlanoPreco.nome_plano).all()
    return jsonify([p.to_dict() for p in planos])


@app.route('/api/planos', methods=['POST'])
def create_plano():
    data = request.get_json(force=True, silent=True) or {}
    nome = (data.get('nome_plano') or '').strip().upper()
    valor = data.get('valor_contrato')
    if not nome or valor is None:
        return jsonify({'error': 'nome_plano e valor_contrato são obrigatórios'}), 400
    try:
        valor = float(valor)
    except (ValueError, TypeError):
        return jsonify({'error': 'valor_contrato deve ser numérico'}), 400

    existing = PlanoPreco.query.filter_by(nome_plano=nome).first()
    if existing:
        existing.valor_contrato = valor
        db.session.commit()
        return jsonify(existing.to_dict())

    plano = PlanoPreco(nome_plano=nome, valor_contrato=valor)
    db.session.add(plano)
    db.session.commit()
    return jsonify(plano.to_dict()), 201


@app.route('/api/planos/<int:pid>', methods=['PUT'])
def update_plano(pid):
    plano = PlanoPreco.query.get_or_404(pid)
    data = request.get_json(force=True, silent=True) or {}
    if 'nome_plano' in data:
        plano.nome_plano = data['nome_plano'].strip().upper()
    if 'valor_contrato' in data:
        try:
            plano.valor_contrato = float(data['valor_contrato'])
        except (ValueError, TypeError):
            return jsonify({'error': 'valor_contrato inválido'}), 400
    db.session.commit()
    # Recalcula diferenças das faturas que usam esse plano
    _recalc_fatura_linhas_por_plano(plano.nome_plano, plano.valor_contrato)
    return jsonify(plano.to_dict())


@app.route('/api/planos/<int:pid>', methods=['DELETE'])
def delete_plano(pid):
    plano = PlanoPreco.query.get_or_404(pid)
    db.session.delete(plano)
    db.session.commit()
    return jsonify({'ok': True})


def _recalc_fatura_linhas_por_plano(nome_plano, valor_contrato):
    """Recalcula diferença em todas as FaturaLinha que possuem esse plano."""
    nome_upper = nome_plano.strip().upper()
    linhas = FaturaLinha.query.filter(
        db.func.upper(FaturaLinha.plano) == nome_upper
    ).all()
    for fl in linhas:
        fl.valor_contrato = valor_contrato
        if fl.valor_fatura is not None:
            fl.diferenca = round(fl.valor_fatura - valor_contrato, 2)
            fl.status = 'ok' if abs(fl.diferenca) < 0.01 else 'divergente'
        else:
            fl.diferenca = None
            fl.status = None
    db.session.commit()


# ============================================================
# IMPORTAR FATURA → COMPARAR COM PLANOS
# ============================================================
@app.route('/api/faturas/importar', methods=['POST'])
def importar_fatura():
    """Importa linhas da fatura no banco e compara com planos cadastrados.
    Aceita JSON com {linhas, meta} ou multipart com PDF."""
    data = request.get_json(force=True, silent=True)

    if data:
        linhas_raw = data.get('linhas', [])
        meta = data.get('meta', {})
        filename = meta.get('filename', '')
        competencia = meta.get('competencia', '')
    else:
        return jsonify({'error': 'Envie JSON com {linhas, meta}'}), 400

    if not isinstance(linhas_raw, list) or not linhas_raw:
        return jsonify({'error': 'Nenhuma linha para importar'}), 400

    # Carrega tabela de preços
    planos_dict = {}
    for p in PlanoPreco.query.all():
        planos_dict[p.nome_plano.strip().upper()] = p.valor_contrato

    imported = []
    for item in linhas_raw:
        numero = re.sub(r'\D', '', str(item.get('numero_vivo', '')))
        plano_nome = (item.get('plano') or '').strip()
        valor_fat = item.get('valor_total') or item.get('valor_fatura')

        if not numero:
            continue

        try:
            valor_fat = float(valor_fat) if valor_fat is not None else None
        except (ValueError, TypeError):
            valor_fat = None

        # Busca preço do plano
        plano_upper = plano_nome.upper()
        valor_ctr = planos_dict.get(plano_upper)
        diferenca = None
        status = None
        if valor_fat is not None and valor_ctr is not None:
            diferenca = round(valor_fat - valor_ctr, 2)
            status = 'ok' if abs(diferenca) < 0.01 else 'divergente'

        fl = FaturaLinha(
            numero_vivo=numero,
            plano=plano_nome,
            valor_fatura=valor_fat,
            valor_contrato=valor_ctr,
            diferenca=diferenca,
            status=status,
            arquivo_origem=filename,
            competencia=competencia,
        )
        db.session.add(fl)
        imported.append(fl)

    db.session.commit()
    return jsonify({
        'ok': True,
        'importados': len(imported),
        'linhas': [fl.to_dict() for fl in imported],
    })


@app.route('/api/faturas/linhas', methods=['GET'])
def list_fatura_linhas():
    """Lista todas as linhas de fatura importadas."""
    linhas = FaturaLinha.query.order_by(FaturaLinha.data_importacao.desc()).all()
    return jsonify([fl.to_dict() for fl in linhas])


@app.route('/api/faturas/linhas', methods=['DELETE'])
def clear_fatura_linhas():
    """Limpa todas as linhas importadas."""
    FaturaLinha.query.delete()
    db.session.commit()
    return jsonify({'ok': True})


@app.route('/api/exportar/csv')
def export_csv():
    competencia = request.args.get('competencia', '').strip()
    status_filter = request.args.get('status', '').strip()
    comparacoes = _get_comparacoes_for_export(competencia, status_filter)

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow([
        'Status', 'Linha', 'No Contrato', 'Valor Contratado',
        'Valor Fatura', 'Diferenca R$', 'Diferenca Pct',
        'Competencia', 'Operadora', 'Observacao'
    ])
    for comp in comparacoes:
        writer.writerow([
            STATUS_LABELS.get(comp.status_comparacao, ''),
            comp.linha_telefone or '',
            comp.numero_contrato or '',
            f'{comp.valor_contratado:.2f}' if comp.valor_contratado is not None else '',
            f'{comp.valor_fatura:.2f}' if comp.valor_fatura is not None else '',
            f'{comp.diferenca_valor:.2f}' if comp.diferenca_valor is not None else '',
            f'{comp.diferenca_percentual:.1f}' if comp.diferenca_percentual is not None else '',
            comp.competencia or '',
            comp.operadora or '',
            comp.observacao or '',
        ])

    csv_bytes = io.BytesIO(output.getvalue().encode('utf-8-sig'))
    fname = f'auditoria_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv'
    return send_file(csv_bytes, mimetype='text/csv', as_attachment=True, download_name=fname)


# ============================================================
# HELPERS DE LISTAGEM
# ============================================================
@app.route('/api/competencias')
def list_competencias():
    rows = (db.session.query(Comparacao.competencia)
            .filter(Comparacao.competencia.isnot(None))
            .distinct()
            .order_by(Comparacao.competencia.desc())
            .all())
    return jsonify([r[0] for r in rows])


@app.route('/api/operadoras')
def list_operadoras():
    rows = (db.session.query(Comparacao.operadora)
            .filter(Comparacao.operadora.isnot(None))
            .distinct().all())
    return jsonify([r[0] for r in rows])


# ============================================================
# RESET — ZERAR BANCO DE DADOS
# ============================================================
@app.route('/api/reset-database', methods=['POST'])
def reset_database():
    """Remove TODOS os dados do banco, preservando a estrutura e configurações padrão."""
    data = request.get_json(force=True, silent=True) or {}
    confirmacao = data.get('confirmacao', '')
    if confirmacao != 'CONFIRMAR':
        return jsonify({'error': 'Envie {"confirmacao":"CONFIRMAR"} para zerar o banco.'}), 400

    # Limpar todas as tabelas em ordem segura (FKs)
    Comparacao.query.delete()
    FaturaLinha.query.delete()
    Conta.query.delete()
    Contrato.query.delete()
    Importacao.query.delete()
    PlanoPreco.query.delete()
    CadastroLinha.query.delete()
    CoopernacVoz.query.delete()
    CoopernacDados.query.delete()
    CoopernacResumo.query.delete()
    Configuracao.query.delete()
    db.session.commit()

    # Recriar configurações padrão
    defaults = [
        ('tolerancia_tipo', 'percentual', 'Tipo de tolerância: fixo (R$) ou percentual (%)'),
        ('tolerancia_valor', '5.0', 'Valor da tolerância'),
        ('apenas_recorrentes', 'false', 'Considerar apenas cobranças recorrentes'),
        ('ignorar_extras', 'false', 'Ignorar taxas extras e multas'),
        ('empresa_nome', 'Auditoria Telecom', 'Nome da empresa'),
    ]
    for chave, valor, descricao in defaults:
        db.session.add(Configuracao(chave=chave, valor=valor, descricao=descricao))
    db.session.commit()

    return jsonify({'ok': True, 'message': 'Banco de dados zerado com sucesso.'})


# ============================================================
# ALERTAS DE VENCIMENTO
# ============================================================
@app.route('/api/alertas/vencimento')
def alertas_vencimento():
    """Retorna faturas que estão vencendo nos próximos dias."""
    from datetime import timedelta
    hoje = datetime.now()
    dia_hoje = hoje.day

    # Busca cadastros com campo vencimento definido
    cadastros = CadastroLinha.query.filter(
        CadastroLinha.vencimento.isnot(None)
    ).all()

    alertas = []
    for c in cadastros:
        try:
            dia_venc = int(c.vencimento)
        except (ValueError, TypeError):
            continue

        # Calcula quantos dias faltam para o vencimento neste mês
        dias_faltam = dia_venc - dia_hoje
        if dias_faltam < 0:
            # Já venceu este mês — calcular para próximo mês
            import calendar
            dias_no_mes = calendar.monthrange(hoje.year, hoje.month)[1]
            dias_faltam = (dias_no_mes - dia_hoje) + dia_venc

        urgencia = 'normal'
        if dias_faltam <= 0:
            urgencia = 'vencido'
        elif dias_faltam <= 3:
            urgencia = 'critico'
        elif dias_faltam <= 7:
            urgencia = 'alerta'

        if dias_faltam <= 10:  # só alerta para vencimentos próximos
            alertas.append({
                'id': c.id,
                'numero_telefone': c.numero_telefone,
                'nome_funcionario': c.nome_funcionario,
                'empresa': c.empresa,
                'vencimento_dia': dia_venc,
                'dias_faltam': dias_faltam,
                'urgencia': urgencia,
                'valor_plano': c.valor_plano,
            })

    alertas.sort(key=lambda a: a['dias_faltam'])
    return jsonify(alertas)


# ============================================================
# CADASTRO COMPLETO DE LINHAS
# ============================================================
@app.route('/api/cadastro', methods=['GET'])
def list_cadastro():
    search = request.args.get('search', '').strip()
    empresa = request.args.get('empresa', '').strip()
    status_linha = request.args.get('status_linha', '').strip()
    query = CadastroLinha.query

    if search:
        like = f'%{search}%'
        query = query.filter(db.or_(
            CadastroLinha.numero_telefone.ilike(like),
            CadastroLinha.nome_funcionario.ilike(like),
            CadastroLinha.matricula_funcionario.ilike(like),
            CadastroLinha.plano.ilike(like),
        ))
    if empresa:
        query = query.filter(CadastroLinha.empresa.ilike(f'%{empresa}%'))
    if status_linha:
        query = query.filter(CadastroLinha.status_linha == status_linha)

    items = query.order_by(CadastroLinha.nome_funcionario).all()
    return jsonify([c.to_dict() for c in items])


@app.route('/api/cadastro', methods=['POST'])
def create_cadastro():
    data = request.get_json(force=True, silent=True) or {}
    numero = re.sub(r'\D', '', str(data.get('numero_telefone', '')))
    if not numero:
        return jsonify({'error': 'Número do telefone é obrigatório'}), 400

    # Buscar valor do contrato a partir do plano cadastrado em PlanoPreco
    plano_nome = (data.get('plano') or '').strip()
    valor_contrato = data.get('valor_contrato')
    valor_fatura = data.get('valor_fatura')

    if valor_contrato is None and plano_nome:
        pp = PlanoPreco.query.filter(db.func.upper(PlanoPreco.nome_plano) == plano_nome.upper()).first()
        if pp:
            valor_contrato = pp.valor_contrato

    # Calcular diferença e conferência
    diferenca = None
    conferencia = 'pendente'
    try:
        vf = float(valor_fatura) if valor_fatura is not None else None
        vc = float(valor_contrato) if valor_contrato is not None else None
        if vf is not None and vc is not None:
            diferenca = round(vf - vc, 2)
            conferencia = 'ok' if abs(diferenca) < 0.01 else 'divergente'
    except (ValueError, TypeError):
        pass

    c = CadastroLinha(
        numero_telefone=numero,
        operadora=(data.get('operadora') or '').strip() or None,
        vencimento=int(data['vencimento']) if data.get('vencimento') else None,
        numero_conta=(data.get('numero_conta') or '').strip() or None,
        matricula_funcionario=(data.get('matricula_funcionario') or '').strip() or None,
        nome_funcionario=(data.get('nome_funcionario') or '').strip() or None,
        centro_custo=(data.get('centro_custo') or '').strip() or None,
        nome_centro_custo=(data.get('nome_centro_custo') or '').strip() or None,
        plano=plano_nome or None,
        valor_plano=float(data['valor_plano']) if data.get('valor_plano') else None,
        empresa=(data.get('empresa') or '').strip() or None,
        conferencia=data.get('conferencia') or conferencia,
        status_linha=data.get('status_linha', 'em_uso'),
        valor_contrato=float(valor_contrato) if valor_contrato is not None else None,
        valor_fatura=float(valor_fatura) if valor_fatura is not None else None,
        diferenca=diferenca,
    )
    db.session.add(c)
    db.session.commit()
    return jsonify(c.to_dict()), 201


@app.route('/api/cadastro/<int:cid>', methods=['PUT'])
def update_cadastro(cid):
    c = CadastroLinha.query.get_or_404(cid)
    data = request.get_json(force=True, silent=True) or {}

    for field in ['numero_telefone', 'operadora', 'numero_conta', 'matricula_funcionario',
                  'nome_funcionario', 'centro_custo', 'nome_centro_custo', 'plano', 'empresa']:
        if field in data:
            val = (data[field] or '').strip() if isinstance(data.get(field), str) else data.get(field)
            if field == 'numero_telefone' and val:
                val = re.sub(r'\D', '', str(val))
            setattr(c, field, val or None)

    if 'vencimento' in data:
        c.vencimento = int(data['vencimento']) if data['vencimento'] else None
    for fld in ['valor_plano', 'valor_contrato', 'valor_fatura']:
        if fld in data:
            try:
                setattr(c, fld, float(data[fld]) if data[fld] is not None else None)
            except (ValueError, TypeError):
                pass
    if 'conferencia' in data:
        c.conferencia = data['conferencia']
    if 'status_linha' in data:
        c.status_linha = data['status_linha']

    # Auto-buscar valor_contrato pelo plano
    if c.valor_contrato is None and c.plano:
        pp = PlanoPreco.query.filter(db.func.upper(PlanoPreco.nome_plano) == c.plano.upper()).first()
        if pp:
            c.valor_contrato = pp.valor_contrato

    # Recalcular diferença
    if c.valor_fatura is not None and c.valor_contrato is not None:
        c.diferenca = round(c.valor_fatura - c.valor_contrato, 2)
        if 'conferencia' not in data:  # só auto-set se não veio manual
            c.conferencia = 'ok' if abs(c.diferenca) < 0.01 else 'divergente'
    else:
        c.diferenca = None

    db.session.commit()
    return jsonify(c.to_dict())


@app.route('/api/cadastro/<int:cid>', methods=['DELETE'])
def delete_cadastro(cid):
    c = CadastroLinha.query.get_or_404(cid)
    db.session.delete(c)
    db.session.commit()
    return jsonify({'ok': True})


@app.route('/api/cadastro/empresas')
def list_empresas():
    """Lista empresas/unidades distintas do cadastro."""
    rows = db.session.query(CadastroLinha.empresa).filter(
        CadastroLinha.empresa.isnot(None)
    ).distinct().order_by(CadastroLinha.empresa).all()
    return jsonify([r[0] for r in rows])


@app.route('/api/cadastro/comparar', methods=['POST'])
def comparar_cadastro():
    """Compara todas as linhas do cadastro com os valores da fatura importada (FaturaLinha).
    Atualiza valor_fatura, diferença e conferência no CadastroLinha."""
    cadastros = CadastroLinha.query.all()
    if not cadastros:
        return jsonify({'error': 'Nenhum cadastro encontrado'}), 400

    # Montar lookup de fatura por número normalizado
    faturas = FaturaLinha.query.all()
    fat_map = {}
    for fl in faturas:
        num = re.sub(r'\D', '', fl.numero_vivo or '')
        if num:
            fat_map[num] = fl

    # Montar lookup de planos
    planos = PlanoPreco.query.all()
    plano_map = {}
    for p in planos:
        plano_map[p.nome_plano.strip().upper()] = p.valor_contrato

    atualizados = 0
    divergentes = 0
    for c in cadastros:
        num = re.sub(r'\D', '', c.numero_telefone or '')
        fl = fat_map.get(num)

        # Atualizar valor_contrato pelo plano se não definido
        if c.plano and not c.valor_contrato:
            vc = plano_map.get(c.plano.strip().upper())
            if vc:
                c.valor_contrato = vc

        if fl and fl.valor_fatura is not None:
            c.valor_fatura = fl.valor_fatura
            if c.valor_contrato is not None:
                c.diferenca = round(c.valor_fatura - c.valor_contrato, 2)
                c.conferencia = 'ok' if abs(c.diferenca) < 0.01 else 'divergente'
                if c.conferencia == 'divergente':
                    divergentes += 1
            atualizados += 1
        elif c.valor_contrato is not None and c.valor_fatura is not None:
            c.diferenca = round(c.valor_fatura - c.valor_contrato, 2)
            c.conferencia = 'ok' if abs(c.diferenca) < 0.01 else 'divergente'
            if c.conferencia == 'divergente':
                divergentes += 1
            atualizados += 1

    db.session.commit()
    return jsonify({
        'ok': True,
        'atualizados': atualizados,
        'divergentes': divergentes,
        'total': len(cadastros),
    })


# ============================================================
# COOPERNAC
# ============================================================
@app.route('/api/coopernac/voz', methods=['GET'])
def list_coopernac_voz():
    items = CoopernacVoz.query.order_by(CoopernacVoz.id).all()
    return jsonify([i.to_dict() for i in items])


@app.route('/api/coopernac/voz', methods=['POST'])
def create_coopernac_voz():
    data = request.get_json(force=True, silent=True) or {}
    item = CoopernacVoz(
        numero=(data.get('numero') or '').strip(),
        descricao=(data.get('descricao') or '').strip(),
        total=float(data.get('total', 0) or 0),
    )
    db.session.add(item)
    db.session.commit()
    return jsonify(item.to_dict()), 201


@app.route('/api/coopernac/voz/<int:vid>', methods=['PUT'])
def update_coopernac_voz(vid):
    item = CoopernacVoz.query.get_or_404(vid)
    data = request.get_json(force=True, silent=True) or {}
    if 'numero' in data: item.numero = (data['numero'] or '').strip()
    if 'descricao' in data: item.descricao = (data['descricao'] or '').strip()
    if 'total' in data: item.total = float(data.get('total', 0) or 0)
    db.session.commit()
    return jsonify(item.to_dict())


@app.route('/api/coopernac/voz/<int:vid>', methods=['DELETE'])
def delete_coopernac_voz(vid):
    item = CoopernacVoz.query.get_or_404(vid)
    db.session.delete(item)
    db.session.commit()
    return jsonify({'ok': True})


@app.route('/api/coopernac/dados', methods=['GET'])
def list_coopernac_dados():
    items = CoopernacDados.query.order_by(CoopernacDados.id).all()
    return jsonify([i.to_dict() for i in items])


@app.route('/api/coopernac/dados', methods=['POST'])
def create_coopernac_dados():
    data = request.get_json(force=True, silent=True) or {}
    item = CoopernacDados(
        numero=(data.get('numero') or '').strip(),
        descricao=(data.get('descricao') or '').strip(),
        total=float(data.get('total', 0) or 0),
    )
    db.session.add(item)
    db.session.commit()
    return jsonify(item.to_dict()), 201


@app.route('/api/coopernac/dados/<int:did>', methods=['PUT'])
def update_coopernac_dados(did):
    item = CoopernacDados.query.get_or_404(did)
    data = request.get_json(force=True, silent=True) or {}
    if 'numero' in data: item.numero = (data['numero'] or '').strip()
    if 'descricao' in data: item.descricao = (data['descricao'] or '').strip()
    if 'total' in data: item.total = float(data.get('total', 0) or 0)
    db.session.commit()
    return jsonify(item.to_dict())


@app.route('/api/coopernac/dados/<int:did>', methods=['DELETE'])
def delete_coopernac_dados(did):
    item = CoopernacDados.query.get_or_404(did)
    db.session.delete(item)
    db.session.commit()
    return jsonify({'ok': True})


@app.route('/api/coopernac/resumo', methods=['GET'])
def list_coopernac_resumo():
    items = CoopernacResumo.query.order_by(CoopernacResumo.id).all()
    return jsonify([i.to_dict() for i in items])


@app.route('/api/coopernac/resumo', methods=['POST'])
def create_coopernac_resumo():
    data = request.get_json(force=True, silent=True) or {}
    item = CoopernacResumo(
        valores=float(data.get('valores', 0) or 0),
        descricao=(data.get('descricao') or '').strip(),
        observacao=(data.get('observacao') or '').strip(),
    )
    db.session.add(item)
    db.session.commit()
    return jsonify(item.to_dict()), 201


@app.route('/api/coopernac/resumo/<int:rid>', methods=['PUT'])
def update_coopernac_resumo(rid):
    item = CoopernacResumo.query.get_or_404(rid)
    data = request.get_json(force=True, silent=True) or {}
    if 'valores' in data: item.valores = float(data.get('valores', 0) or 0)
    if 'descricao' in data: item.descricao = (data['descricao'] or '').strip()
    if 'observacao' in data: item.observacao = (data['observacao'] or '').strip()
    db.session.commit()
    return jsonify(item.to_dict())


@app.route('/api/coopernac/resumo/<int:rid>', methods=['DELETE'])
def delete_coopernac_resumo(rid):
    item = CoopernacResumo.query.get_or_404(rid)
    db.session.delete(item)
    db.session.commit()
    return jsonify({'ok': True})


@app.route('/api/relatorios/fatura-linhas')
def relatorio_fatura_linhas():
    """Retorna linhas de fatura com filtros de data e número."""
    data_inicio = request.args.get('data_inicio', '').strip()
    data_fim = request.args.get('data_fim', '').strip()
    numero = request.args.get('numero', '').strip()

    query = FaturaLinha.query
    if data_inicio:
        try:
            dt = datetime.strptime(data_inicio, '%Y-%m-%d')
            query = query.filter(FaturaLinha.data_importacao >= dt)
        except ValueError:
            pass
    if data_fim:
        try:
            dt = datetime.strptime(data_fim, '%Y-%m-%d') + timedelta(days=1)
            query = query.filter(FaturaLinha.data_importacao < dt)
        except ValueError:
            pass
    if numero:
        like = f'%{re.sub(r"[^0-9]", "", numero)}%'
        query = query.filter(FaturaLinha.numero_vivo.ilike(like))

    items = query.order_by(FaturaLinha.data_importacao.desc()).all()
    return jsonify([fl.to_dict() for fl in items])


@app.route('/api/exportar/fatura-excel')
def export_fatura_excel():
    """Exporta linhas de fatura filtradas para Excel."""
    data_inicio = request.args.get('data_inicio', '').strip()
    data_fim = request.args.get('data_fim', '').strip()
    numero = request.args.get('numero', '').strip()

    query = FaturaLinha.query
    if data_inicio:
        try:
            dt = datetime.strptime(data_inicio, '%Y-%m-%d')
            query = query.filter(FaturaLinha.data_importacao >= dt)
        except ValueError:
            pass
    if data_fim:
        try:
            dt = datetime.strptime(data_fim, '%Y-%m-%d') + timedelta(days=1)
            query = query.filter(FaturaLinha.data_importacao < dt)
        except ValueError:
            pass
    if numero:
        like = f'%{re.sub(r"[^0-9]", "", numero)}%'
        query = query.filter(FaturaLinha.numero_vivo.ilike(like))

    items = query.order_by(FaturaLinha.data_importacao.desc()).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Faturas'
    ws.freeze_panes = 'A2'

    headers = ['Número', 'Plano', 'Valor Fatura', 'Valor Contrato', 'Diferença', 'Status', 'Competência', 'Arquivo', 'Data Importação']
    header_fill = PatternFill('solid', fgColor='660099')
    header_font = Font(color='FFFFFF', bold=True, size=10)
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    for ri, fl in enumerate(items, 2):
        ws.cell(row=ri, column=1, value=fl.numero_vivo)
        ws.cell(row=ri, column=2, value=fl.plano)
        ws.cell(row=ri, column=3, value=fl.valor_fatura)
        ws.cell(row=ri, column=4, value=fl.valor_contrato)
        ws.cell(row=ri, column=5, value=fl.diferenca)
        ws.cell(row=ri, column=6, value=fl.status or '')
        ws.cell(row=ri, column=7, value=fl.competencia or '')
        ws.cell(row=ri, column=8, value=fl.arquivo_origem or '')
        ws.cell(row=ri, column=9, value=fl.data_importacao.strftime('%d/%m/%Y %H:%M') if fl.data_importacao else '')

        # Colorir status
        if fl.status == 'divergente':
            for c in range(1, 10):
                ws.cell(row=ri, column=c).fill = PatternFill('solid', fgColor='FFCCCC')
        elif fl.status == 'ok':
            for c in range(1, 10):
                ws.cell(row=ri, column=c).fill = PatternFill('solid', fgColor='CCFFCC')

    for i, w in enumerate([18, 35, 15, 15, 12, 12, 12, 30, 18], 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name=f'faturas_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx')


# ============================================================
# ENTRY POINT
# ============================================================
if __name__ == '__main__':
    init_db()
    print('\n' + '=' * 60)
    print('  Auditoria Telecom - Sistema iniciado com sucesso!')
    print('  Acesse: http://localhost:5000')
    print('=' * 60 + '\n')
    app.run(debug=True, host='0.0.0.0', port=5000, threaded=True)
